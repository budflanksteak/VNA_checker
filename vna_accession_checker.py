#!/usr/bin/env python3
"""
vna_accession_checker.py

Read an Excel file containing a list of accession numbers, query a DICOM VNA
(Vendor Neutral Archive) by accession number using a DIMSE C-FIND, determine
whether images are stored for each accession and which modalities are present,
then write the results back into the same Excel workbook.

The query is a STUDY-level C-FIND against the Study Root Query/Retrieve
Information Model. For each accession number the script reports:

    * Images Present   -> Yes / No / Error
    * Modality         -> comma-separated list of modalities (e.g. "CT, MR")
    * Study Count      -> number of studies matching the accession number
    * Instance Count   -> total DICOM instances across matching studies
    * Query Notes      -> status / error detail

Connection details (VNA AE title, host, port and this application's calling
AE title) are supplied via command-line arguments or environment variables.

Example
-------
    python vna_accession_checker.py accessions.xlsx \
        --host 10.0.0.5 --port 104 \
        --called-aet VNA_SCP --calling-aet VNA_CHECK

Environment variable equivalents: VNA_HOST, VNA_PORT, VNA_CALLED_AET,
VNA_CALLING_AET.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from dataclasses import dataclass, field
from typing import Iterable, Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - dependency guard
    sys.exit(
        "Missing dependency 'openpyxl'. Install requirements first:\n"
        "    pip install -r requirements.txt"
    )

try:
    from pydicom.dataset import Dataset
    from pynetdicom import AE
    from pynetdicom.sop_class import (
        StudyRootQueryRetrieveInformationModelFind,
    )
except ImportError:  # pragma: no cover - dependency guard
    sys.exit(
        "Missing dependency 'pynetdicom'/'pydicom'. Install requirements first:\n"
        "    pip install -r requirements.txt"
    )


logger = logging.getLogger("vna_accession_checker")


# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
# Column headers written back into the workbook. Order is preserved.
RESULT_COLUMNS = [
    "Images Present",
    "Modality",
    "Study Count",
    "Instance Count",
    "Query Notes",
]

# Header names (lower-cased) that identify the accession-number column.
ACCESSION_HEADER_CANDIDATES = {
    "accession number",
    "accession",
    "accession no",
    "accession no.",
    "accession #",
    "accession_number",
    "accessionnumber",
    "accession num",
    "accnum",
    "acc num",
    "acc #",
    "acc",
}


@dataclass
class VNAConfig:
    """DICOM connection parameters for the remote VNA (the C-FIND SCP)."""

    host: str
    port: int
    called_aet: str
    calling_aet: str
    connection_timeout: float = 30.0
    query_series_fallback: bool = True


@dataclass
class QueryResult:
    """Aggregated result of querying one accession number."""

    images_present: str = "No"          # Yes / No / Error
    modalities: list[str] = field(default_factory=list)
    study_count: int = 0
    instance_count: int = 0
    notes: str = ""

    @property
    def modality_str(self) -> str:
        return ", ".join(self.modalities)


# --------------------------------------------------------------------------- #
# DICOM querying
# --------------------------------------------------------------------------- #
class VNAClient:
    """Thin wrapper around a pynetdicom association to a DICOM VNA."""

    def __init__(self, config: VNAConfig) -> None:
        self.config = config
        self.ae = AE(ae_title=config.calling_aet)
        self.ae.add_requested_context(StudyRootQueryRetrieveInformationModelFind)
        self.ae.acse_timeout = config.connection_timeout
        self.ae.dimse_timeout = config.connection_timeout
        self.ae.network_timeout = config.connection_timeout

    def verify(self) -> None:
        """Open and immediately release an association to confirm connectivity.

        Raises ConnectionError if the VNA cannot be reached or rejects us.
        """
        assoc = self._associate()
        assoc.release()

    def _associate(self):
        assoc = self.ae.associate(
            self.config.host,
            self.config.port,
            ae_title=self.config.called_aet,
        )
        if not assoc.is_established:
            raise ConnectionError(
                f"Association rejected/aborted by {self.config.called_aet} "
                f"at {self.config.host}:{self.config.port}"
            )
        return assoc

    def query_accession(self, accession: str) -> QueryResult:
        """Perform a STUDY-level C-FIND for one accession number."""
        result = QueryResult()

        identifier = Dataset()
        identifier.QueryRetrieveLevel = "STUDY"
        identifier.AccessionNumber = accession
        # Return keys (universal matching -> request these attributes back).
        identifier.StudyInstanceUID = ""
        identifier.ModalitiesInStudy = ""
        identifier.NumberOfStudyRelatedInstances = ""
        identifier.StudyDate = ""
        identifier.StudyDescription = ""

        modalities: set[str] = set()
        studies_needing_series_lookup: list[str] = []

        try:
            assoc = self._associate()
        except ConnectionError as exc:
            result.images_present = "Error"
            result.notes = str(exc)
            return result

        try:
            responses = assoc.send_c_find(
                identifier, StudyRootQueryRetrieveInformationModelFind
            )
            for status, ds in responses:
                if status is None:
                    result.images_present = "Error"
                    result.notes = "No response / connection timed out during C-FIND"
                    break

                # 0xFF00 / 0xFF01 = Pending (a matching record follows).
                if status.Status in (0xFF00, 0xFF01):
                    if ds is None:
                        continue
                    result.study_count += 1

                    n_instances = getattr(ds, "NumberOfStudyRelatedInstances", None)
                    if n_instances not in (None, ""):
                        try:
                            result.instance_count += int(n_instances)
                        except (TypeError, ValueError):
                            pass

                    ds_modalities = getattr(ds, "ModalitiesInStudy", None)
                    if ds_modalities:
                        # May be a single value or a MultiValue of strings.
                        if isinstance(ds_modalities, str):
                            modalities.add(ds_modalities)
                        else:
                            modalities.update(str(m) for m in ds_modalities)
                    elif self.config.query_series_fallback:
                        study_uid = getattr(ds, "StudyInstanceUID", None)
                        if study_uid:
                            studies_needing_series_lookup.append(str(study_uid))

                elif status.Status == 0x0000:
                    # Success -> matching complete.
                    break
                else:
                    # Any failure/cancel/warning status.
                    result.images_present = "Error"
                    result.notes = (
                        f"C-FIND returned status 0x{status.Status:04X}"
                    )
                    break
        finally:
            assoc.release()

        # Fill in modalities for studies that did not report ModalitiesInStudy.
        if studies_needing_series_lookup and "Error" not in result.images_present:
            for study_uid in studies_needing_series_lookup:
                modalities.update(self._series_modalities(study_uid))

        if result.images_present != "Error":
            result.modalities = sorted(modalities)
            if result.study_count > 0:
                result.images_present = "Yes"
                if not result.notes:
                    result.notes = "OK"
            else:
                result.images_present = "No"
                if not result.notes:
                    result.notes = "No matching studies in VNA"

        return result

    def _series_modalities(self, study_instance_uid: str) -> set[str]:
        """SERIES-level C-FIND to collect Modality when ModalitiesInStudy is absent."""
        modalities: set[str] = set()
        identifier = Dataset()
        identifier.QueryRetrieveLevel = "SERIES"
        identifier.StudyInstanceUID = study_instance_uid
        identifier.SeriesInstanceUID = ""
        identifier.Modality = ""

        try:
            assoc = self._associate()
        except ConnectionError:
            return modalities

        try:
            responses = assoc.send_c_find(
                identifier, StudyRootQueryRetrieveInformationModelFind
            )
            for status, ds in responses:
                if status and status.Status in (0xFF00, 0xFF01) and ds is not None:
                    modality = getattr(ds, "Modality", None)
                    if modality:
                        modalities.add(str(modality))
        finally:
            assoc.release()

        return modalities


# --------------------------------------------------------------------------- #
# Excel handling
# --------------------------------------------------------------------------- #
def find_accession_column(worksheet, explicit: Optional[str]) -> int:
    """Return the 1-based column index that holds accession numbers.

    If `explicit` is given it may be a header name or a column letter/number.
    Otherwise the header row (row 1) is scanned for a known accession header.
    """
    header_row = 1

    if explicit:
        # Column letter (e.g. "B")?
        if explicit.isalpha() and len(explicit) <= 3:
            try:
                return openpyxl.utils.column_index_from_string(explicit.upper())
            except ValueError:
                pass
        # Numeric column index?
        if explicit.isdigit():
            return int(explicit)
        # Otherwise treat as a header name.
        for cell in worksheet[header_row]:
            if cell.value and str(cell.value).strip().lower() == explicit.strip().lower():
                return cell.column
        raise ValueError(f"Could not find a column matching '{explicit}'")

    for cell in worksheet[header_row]:
        if cell.value and str(cell.value).strip().lower() in ACCESSION_HEADER_CANDIDATES:
            return cell.column

    raise ValueError(
        "Could not auto-detect the accession-number column. "
        "Pass --accession-column with the header name or column letter."
    )


def ensure_result_columns(worksheet) -> dict[str, int]:
    """Ensure the result columns exist in the header row; return name->col index.

    Reuses columns that already carry the expected headers (so re-runs update
    in place instead of appending duplicates).
    """
    header_row = 1
    existing: dict[str, int] = {}
    max_col = worksheet.max_column
    for col in range(1, max_col + 1):
        value = worksheet.cell(row=header_row, column=col).value
        if value is not None:
            existing[str(value).strip()] = col

    columns: dict[str, int] = {}
    next_col = max_col + 1
    for name in RESULT_COLUMNS:
        if name in existing:
            columns[name] = existing[name]
        else:
            worksheet.cell(row=header_row, column=next_col, value=name)
            columns[name] = next_col
            next_col += 1
    return columns


def iter_accession_rows(worksheet, accession_col: int) -> Iterable[tuple[int, str]]:
    """Yield (row_index, accession_number) for each data row with a value."""
    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=accession_col).value
        if value is None:
            continue
        accession = str(value).strip()
        if accession:
            yield row, accession


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def process_workbook(
    excel_path: str,
    config: VNAConfig,
    accession_column: Optional[str],
    sheet_name: Optional[str],
) -> None:
    logger.info("Opening workbook: %s", excel_path)
    workbook = openpyxl.load_workbook(excel_path)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    accession_col = find_accession_column(worksheet, accession_column)
    logger.info(
        "Using accession-number column %s (%s)",
        get_column_letter(accession_col),
        worksheet.cell(row=1, column=accession_col).value,
    )

    result_cols = ensure_result_columns(worksheet)

    rows = list(iter_accession_rows(worksheet, accession_col))
    if not rows:
        logger.warning("No accession numbers found in the workbook.")
        return

    logger.info("Found %d accession number(s) to query.", len(rows))

    client = VNAClient(config)
    logger.info(
        "Verifying connectivity to %s@%s:%d ...",
        config.called_aet,
        config.host,
        config.port,
    )
    client.verify()
    logger.info("Connection OK.")

    for idx, (row, accession) in enumerate(rows, start=1):
        logger.info("[%d/%d] Querying accession %s", idx, len(rows), accession)
        result = client.query_accession(accession)

        worksheet.cell(row=row, column=result_cols["Images Present"],
                       value=result.images_present)
        worksheet.cell(row=row, column=result_cols["Modality"],
                       value=result.modality_str)
        worksheet.cell(row=row, column=result_cols["Study Count"],
                       value=result.study_count)
        worksheet.cell(row=row, column=result_cols["Instance Count"],
                       value=result.instance_count)
        worksheet.cell(row=row, column=result_cols["Query Notes"],
                       value=result.notes)

        logger.info(
            "    -> %s | modalities=%s | studies=%d | instances=%d",
            result.images_present,
            result.modality_str or "-",
            result.study_count,
            result.instance_count,
        )

    logger.info("Saving results back into %s", excel_path)
    workbook.save(excel_path)
    logger.info("Done.")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Query a DICOM VNA by accession number (DIMSE C-FIND) and write "
            "image-presence + modality results back into the Excel file."
        )
    )
    parser.add_argument("excel_file", help="Path to the .xlsx workbook.")
    parser.add_argument(
        "--host",
        default=os.environ.get("VNA_HOST"),
        help="VNA hostname/IP (env: VNA_HOST).",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=int(os.environ["VNA_PORT"]) if os.environ.get("VNA_PORT") else None,
        help="VNA DICOM port, e.g. 104 (env: VNA_PORT).",
    )
    parser.add_argument(
        "--called-aet",
        default=os.environ.get("VNA_CALLED_AET"),
        help="The VNA's AE Title (called AE) (env: VNA_CALLED_AET).",
    )
    parser.add_argument(
        "--calling-aet",
        default=os.environ.get("VNA_CALLING_AET", "VNA_CHECK"),
        help="This application's AE Title (calling AE) (env: VNA_CALLING_AET).",
    )
    parser.add_argument(
        "--accession-column",
        default=None,
        help=(
            "Header name, column letter (e.g. B) or number for the accession "
            "column. Auto-detected from the header row if omitted."
        ),
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name to process (defaults to the active sheet).",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=30.0,
        help="DICOM network/association/DIMSE timeout in seconds (default 30).",
    )
    parser.add_argument(
        "--no-series-fallback",
        action="store_true",
        help=(
            "Do not fall back to a SERIES-level query when the VNA omits "
            "ModalitiesInStudy."
        ),
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Enable debug logging (includes pynetdicom protocol logs).",
    )
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
        datefmt="%H:%M:%S",
    )
    if not args.verbose:
        # Quiet the fairly chatty pynetdicom logger unless -v is set.
        logging.getLogger("pynetdicom").setLevel(logging.WARNING)

    missing = [
        name
        for name, value in (
            ("--host / VNA_HOST", args.host),
            ("--port / VNA_PORT", args.port),
            ("--called-aet / VNA_CALLED_AET", args.called_aet),
        )
        if not value
    ]
    if missing:
        parser.error("Missing required connection settings: " + ", ".join(missing))

    if not os.path.isfile(args.excel_file):
        parser.error(f"Excel file not found: {args.excel_file}")

    config = VNAConfig(
        host=args.host,
        port=args.port,
        called_aet=args.called_aet,
        calling_aet=args.calling_aet,
        connection_timeout=args.timeout,
        query_series_fallback=not args.no_series_fallback,
    )

    try:
        process_workbook(
            excel_path=args.excel_file,
            config=config,
            accession_column=args.accession_column,
            sheet_name=args.sheet,
        )
    except (ValueError, ConnectionError) as exc:
        logger.error("%s", exc)
        return 1
    except KeyboardInterrupt:  # pragma: no cover
        logger.error("Interrupted.")
        return 130

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
